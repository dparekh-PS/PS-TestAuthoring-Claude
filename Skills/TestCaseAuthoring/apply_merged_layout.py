#!/usr/bin/env python3
"""
Apply the STANDARD test-case presentation (EXCEL_SPECIFICATION v2.5): merge the TC-level
columns vertically across each test case's step rows, so Test Case ID, Requirement Title,
Test Case Title, Pre-Conditions, and Priority appear ONCE per test case (value in the top
row), while Step#, Test Step, and Expected Result vary per row. Same 8 columns.

Also removes prior row-grouping/forward-fill artefacts, RIGHT-SIZES row heights to the
actual per-row content (no more tall empty gaps), and re-stamps the schema version
(read from validate_workbook.py — the single source of truth).

Usage: python apply_merged_layout.py <workbook.xlsx> [...]
"""
import sys, os, re, math, importlib.util, openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _schema_version():
    """Single source of truth: read SCHEMA_VERSION from the validator beside this script."""
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validate_workbook.py")
    spec = importlib.util.spec_from_file_location("vw_schema", p)
    m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
    return m.SCHEMA_VERSION

MERGE_COLS = [1, 2, 3, 4, 8]          # A ID, B Requirement, C Title, D Pre-Conditions, H Priority
STEP_COLS  = [6, 7]                    # F Test Step, G Expected Result (drive per-row height)
LINE_PT = 15.0                         # approx points per wrapped text line
TOPWRAP = Alignment(vertical="top", wrap_text=True)
med = Side(style="medium", color="B4C6E7")

def _norm(v): return "" if v is None else str(v).strip()

def _feature_sheets(wb):
    out = []
    for ws in wb.worksheets:
        if ws.title in ("Master Summary", "Review Summary"):
            continue
        for r in range(1, 9):
            if _norm(ws.cell(row=r, column=1).value) == "Test Case ID":
                out.append((ws, r)); break
    return out

def _lines(text, width):
    """Estimate wrapped lines for `text` in a column `width` chars wide."""
    if text is None or str(text).strip() == "":
        return 1
    cpl = max(8, int(width) - 1)
    total = 0
    for seg in str(text).split("\n"):
        seg = seg.rstrip()
        total += max(1, math.ceil(len(seg) / cpl)) if seg else 1
    return max(1, total)

def apply(path):
    wb = openpyxl.load_workbook(path)
    tcs = 0
    for ws, hrow in _feature_sheets(wb):
        widths = {c: (ws.column_dimensions[get_column_letter(c)].width or 10) for c in range(1, 9)}

        # 0) IDEMPOTENCY: undo any existing per-TC vertical merges so a re-run (regeneration)
        #    doesn't crash writing to read-only MergedCell continuation cells. Only unmerge
        #    ranges below the header row — the title/source banners (rows <= hrow) are left
        #    intact. openpyxl keeps each merge's top-left value on unmerge.
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row > hrow:
                ws.unmerge_cells(str(rng))

        # 1) strip grouping + reset any baked row heights (they'll be recomputed)
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].hidden = False
            ws.row_dimensions[r].outline_level = 0
            if r > hrow:
                ws.row_dimensions[r].height = None
        ws.sheet_properties.outlinePr.summaryBelow = True

        # 2) find TC groups = runs of equal, non-empty col-A value (forward-filled input)
        runs = []; start = None; cur = None; prev = None
        for r in range(hrow + 1, ws.max_row + 1):
            a = _norm(ws.cell(row=r, column=1).value)
            blank = (a == "" and _norm(ws.cell(row=r, column=6).value) == "" and ws.cell(row=r, column=5).value is None)
            if blank:
                continue
            if a and a != cur:
                if start is not None: runs.append((start, prev, cur))
                start = r; cur = a
            prev = r
        if start is not None: runs.append((start, prev, cur))

        # 3) merge TC-level columns; keep value only in the top row
        for s, e, _id in runs:
            tcs += 1
            for c in MERGE_COLS:
                if e > s:
                    for rr in range(s + 1, e + 1):
                        ws.cell(row=rr, column=c).value = None
                    ws.merge_cells(start_row=s, end_row=e, start_column=c, end_column=c)
                ws.cell(row=s, column=c).alignment = TOPWRAP
            for c in range(1, 9):                     # separator under each test case
                cell = ws.cell(row=e, column=c); b = cell.border
                cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=med)

        # 4) right-size row heights: per-row from step cells; ensure merged block fits
        base = {}
        for s, e, _id in runs:
            for r in range(s, e + 1):
                base[r] = max(_lines(ws.cell(row=r, column=c).value, widths[c]) for c in STEP_COLS)
            merged_need = max(
                _lines(ws.cell(row=s, column=3).value, widths[3]),   # Test Case Title
                _lines(ws.cell(row=s, column=4).value, widths[4]),   # Pre-Conditions
                _lines(ws.cell(row=s, column=2).value, widths[2]),   # Requirement Title
            )
            span = sum(base[r] for r in range(s, e + 1))
            if span < merged_need:                    # give the deficit to the first row
                base[s] += (merged_need - span)
        for r, lines in base.items():
            ws.row_dimensions[r].height = round(lines * LINE_PT + 3, 1)

    kw = wb.properties.keywords or ""
    kw = re.sub(r"schema:[0-9.]+", "", kw).strip()
    wb.properties.keywords = (kw + f" schema:{_schema_version()}").strip()
    wb.save(path)
    print(f"{path}: merged {tcs} test case(s), right-sized rows, schema:{_schema_version()}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.stderr.write(__doc__); sys.exit(2)
    for p in sys.argv[1:]:
        apply(p)
