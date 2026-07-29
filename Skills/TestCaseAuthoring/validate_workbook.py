#!/usr/bin/env python3
"""
PS-TestAuthoring — Deterministic Workbook Validator
===================================================

Single machine-readable source of truth for the Excel output contract.
The LLM proposes a workbook; THIS SCRIPT disposes. A workbook is only
"review-ready" if this validator exits 0.

Usage:
    python validate_workbook.py <workbook.xlsx> [<workbook2.xlsx> ...]
    python validate_workbook.py --json <workbook.xlsx>
    python validate_workbook.py --rules      # emit the canonical rule catalog (Markdown)
    python validate_workbook.py --register <workbook.xlsx>   # record a PASSED workbook's TC IDs
                                                             # in the persistent id_ledger.json

Exit codes:
    0  = PASS (no Fatal or Blocking findings)
    1  = FAIL (one or more Fatal or Blocking findings)
    2  = usage / file error

Severity model:
    FATAL     -> workbook is invalid, must not be delivered
    BLOCKING  -> must be fixed before delivery
    WARNING   -> should be fixed; does not block

AUTHORITY (single source of truth):
    * This file is authoritative for the VALIDATION RULE CODES and their SEVERITIES — they
      are declared once in the RULES catalog below. VALIDATION_ENGINE.md renders that catalog
      from `--rules` inside generated markers, and `Skills/lint_docs.py` FAILS if the doc
      table drifts from this file, so the three former copies can no longer disagree.
    * EXCEL_SPECIFICATION.md remains authoritative for the human-readable OUTPUT CONTRACT
      (columns, sheets, IDs, naming); the two must agree. Bump SCHEMA_VERSION whenever the
      column set, ID format, or sheet contract changes.
"""

import sys
import os
import re
import json
from datetime import date

# Shared validator substrate (finding model + severity taxonomy + CLI) lives in Skills/_base,
# so every skill's validator uses the same contract. This workbook validator is one impl.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "_base"))
from validator_base import Report, print_report, run_cli  # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl --break-system-packages\n")
    sys.exit(2)

# --------------------------------------------------------------------------
# CANONICAL CONTRACT  (the single source of truth)
# --------------------------------------------------------------------------

SCHEMA_VERSION = "2.5"

# Exact feature-sheet headers, in order. No columns may be added or removed.
FEATURE_HEADERS = [
    "Test Case ID",
    "Requirement Title",
    "Test Case Title",
    "Pre-Conditions",
    "Step#",
    "Test Step",
    "Expected Result",
    "Priority",
]

FIXED_SHEETS = ["Master Summary", "Review Summary"]  # positions 1 and 2

PRIORITY_VALUES = {"High", "Medium", "Low"}
TITLE_PREFIXES = ("[Positive]", "[Negative]", "[Edge Case]")

# Project+story scoped, globally unique. e.g. SAMP-125-TC-001
TC_ID_RE = re.compile(r"^[A-Z][A-Z0-9]*-\d+-TC-\d{3,}$")   # full-cell match (ID cells)
TC_ID_SCAN = re.compile(r"[A-Z][A-Z0-9]*-\d+-TC-\d{3,}")   # find IDs embedded in text (RTM cells)

MIN_STEPS_PER_TC = 3

# Expected Result quality gate (owned by TEST_CASE_GENERATION.md §6.4). A step whose ENTIRE
# Expected Result is nothing more than one of these bare phrases FAILS. Multi-point results
# (more than one bullet line) are never matched, so rich results are never false-flagged.
BANNED_EXPECTED = {
    "saved", "saved successfully", "record created", "record saved", "record updated",
    "rule created", "rule saved", "rule updated", "validation displayed", "validation shown",
    "quote submitted", "quote saved", "workflow triggered", "workflow executed",
    "email sent", "success message displayed", "notification generated", "notification sent",
    "done", "as expected", "works correctly", "works as expected", "it works as expected",
    "success", "successful", "updated", "created", "submitted", "displayed", "completed",
    "pass", "passed", "ok", "okay", "no error", "no errors", "verified",
}
_BULLET_RE = re.compile(r"^[\s•●▪\-\*]+")
_THIN_FINAL_MAX = 40   # a one-line final-step result shorter than this is likely too thin

def _expected_lines(expected):
    """Non-empty lines of an Expected Result with leading bullet markers stripped."""
    out = []
    for ln in str(expected).split("\n"):
        ln = _BULLET_RE.sub("", ln).strip()
        if ln:
            out.append(ln)
    return out

def _is_banned_expected(expected):
    """True only when the whole cell is a single bare phrase from BANNED_EXPECTED."""
    lines = _expected_lines(expected)
    if len(lines) != 1:
        return False
    t = re.sub(r"\s+", " ", lines[0].strip().strip('"\'').rstrip(".!").strip().lower())
    t = re.sub(r"^(the |a |an )", "", t)
    return t in BANNED_EXPECTED


# --------------------------------------------------------------------------
# CANONICAL RULE CATALOG  (single source of truth for codes + severities)
# --------------------------------------------------------------------------
# Every finding this validator can emit is declared here exactly once. VALIDATION_ENGINE.md
# renders this catalog from `--rules` between generated markers, and lint_docs.py FAILS if the
# doc drifts from it — so codes/severities can never diverge across the docs again. `severity`
# is the primary severity; a couple of checks downgrade to a lesser severity when a cell or
# column cannot be located (noted in the description).
RULES = [
    ("LOAD",  "FATAL",    "Workbook cannot be opened or parsed."),
    ("SV-01", "BLOCKING", "Schema version stamp missing from document Keywords (expected `schema:X.Y`)."),
    ("SV-02", "BLOCKING", "Schema version stamp does not match the validator's SCHEMA_VERSION."),
    ("WV-01", "FATAL",    "First sheet is not 'Master Summary'."),
    ("WV-03", "FATAL",    "No feature worksheet exists after Master Summary."),
    ("WV-04", "FATAL",    "A column beyond the canonical 8-column schema is present."),
    ("WV-05", "FATAL",    "Feature-sheet header row is missing or does not match the exact 8 headers."),
    ("WV-06", "BLOCKING", "Sheet name exceeds 31 characters or contains prohibited characters."),
    ("DV-01", "FATAL",    "A step row appears before any Test Case ID (the first TC row must carry the ID)."),
    ("DV-02", "FATAL",    "Blank Test Case Title on a test case."),
    ("DV-03", "FATAL",    "Blank Test Step on a step row."),
    ("DV-04", "FATAL",    "Blank Expected Result on a step row."),
    ("DV-05", "BLOCKING", "Blank Priority on a test case."),
    ("DV-06", "BLOCKING", "Priority is not exactly High, Medium, or Low."),
    ("DV-07", "WARNING",  "Step# is not a sequential integer starting at 1 per test case."),
    ("DV-08", "WARNING",  "Test Case Title does not start with [Positive], [Negative], or [Edge Case]."),
    ("DV-09", "FATAL",    "A Test Case ID reappears as a new test case (IDs must be globally unique)."),
    ("DV-10", "WARNING",  "Test case has fewer than the minimum number of steps (3)."),
    ("DV-11", "FATAL",    "Test Case ID does not match the required PROJECT-STORY-TC-NNN format."),
    ("DV-12", "BLOCKING", "Blank Requirement Title on a test case."),
    ("DV-13", "BLOCKING", "Blank Pre-Conditions on a test case."),
    ("ER-01", "BLOCKING", "The entire Expected Result is a single bare vague phrase (see BANNED_EXPECTED)."),
    ("ER-02", "WARNING",  "The final (key verification) step carries only a thin one-line Expected Result."),
    ("CV-01", "FATAL",    "Master Summary test-case count does not match the actual feature-sheet count (Warning if the Test Cases column cannot be located)."),
    ("CV-06", "FATAL",    "A Master Summary row reports AC Coverage % below 100% or blank (Blocking when the coverage column/row cannot be parsed)."),
    ("CV-07", "BLOCKING", "Master Summary internal inconsistency (a feature with test cases but 0 acceptance criteria, or a feature row with 0 test cases)."),
    ("CV-08", "FATAL",    "Coverage ledger: an acceptance criterion is uncovered, is covered by a test case not present in the workbook, or is covered by a test case that belongs to a different feature sheet (cross-sheet coverage is not allowed)."),
    ("CV-09", "FATAL",    "Coverage ledger: a feature's AC count (or the grand total / coverage %) does not reconcile with the Master Summary."),
    ("CV-10", "BLOCKING", "Coverage ledger: an acceptance criterion has no source anchor (each AC must cite where in the source it came from)."),
    ("CV-11", "BLOCKING", "No coverage ledger sidecar found (or it is malformed) — AC coverage cannot be verified against a source-anchored AC list; the ledger is a required deliverable."),
    ("NS-01", "FATAL",    "Cross-workbook Test Case ID collision: an ID was already issued in a different workbook (per the persistent id_ledger.json). IDs must be globally unique across workbooks, projects, and business units."),
    ("NS-02", "WARNING",  "A Test Case ID uses a project key that is not an enabled project in project_registry.json — register the project so its IDs are namespaced and governed."),
    ("DP-01", "BLOCKING", "A real-looking email address appears in a cell — reproduce personal data as an environment-independent characteristic/placeholder, never a real value (DATA_HANDLING.md)."),
    ("DP-02", "BLOCKING", "A probable secret/credential (API key, token, private key, JWT) appears in a cell — secrets must never be reproduced in test cases (DATA_HANDLING.md)."),
    ("DUP-01","WARNING",  "Two test cases on the same feature sheet are near-duplicates (identical title, or identical step + expected-result sequence) — merge or differentiate them."),
    ("EI-01", "WARNING",  "A cell contains a probable hard-coded environment/record identifier (e.g. a quote id like Q-100245 or a Salesforce record id) — use an environment-independent characteristic instead."),
    ("ST-01", "WARNING",  "A test step bundles multiple UI actions (a compound step: two action verbs joined by 'and'/'then') — split it so each step performs exactly one action (TEST_CASE_GENERATION.md 5.2)."),
    ("ST-02", "WARNING",  "Suite is under-decomposed: an outsized share of test cases sit at the 3-step minimum (well-authored manual cases usually run 4-8 atomic steps; only simple presence/absence checks belong at the floor) (TEST_CASE_GENERATION.md 5.1)."),
]

RULES_BEGIN = "<!-- RULES:BEGIN generated by validate_workbook.py --rules; do not edit by hand -->"
RULES_END   = "<!-- RULES:END -->"

def emit_rules_markdown():
    """Deterministic Markdown table of the canonical rule catalog."""
    out = ["| Code | Severity | Check |", "|------|----------|-------|"]
    out += [f"| {code} | {sev} | {desc} |" for code, sev, desc in RULES]
    return "\n".join(out)

def emit_rules_block():
    """The generated block (with markers) that the docs embed and the linter checks."""
    return f"{RULES_BEGIN}\n{emit_rules_markdown()}\n{RULES_END}"

def _emitted_codes():
    """Every code the validator can actually emit (scanned from this source file)."""
    src = open(__file__, encoding="utf-8").read()
    return set(re.findall(r'rep\.(?:fatal|blocking|warn)\("([A-Z0-9-]+)"', src))

def rules_consistency_errors():
    """Drift between the RULES catalog and the codes actually emitted (should always be empty)."""
    cataloged = {c for c, _, _ in RULES}
    emitted = _emitted_codes()
    problems  = [f"code {c} is emitted but missing from the RULES catalog" for c in sorted(emitted - cataloged)]
    problems += [f"code {c} is in the RULES catalog but never emitted" for c in sorted(cataloged - emitted)]
    return problems


# Report / print_report / run_cli are imported from Skills/_base/validator_base.py (above).


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def _norm(v):
    return "" if v is None else str(v).strip()

def find_header_row(ws, expected_first="Test Case ID", max_scan=8):
    """Feature-sheet headers may sit on row 3 or 4 depending on the title block.
    Locate the row whose column A equals the first expected header."""
    for r in range(1, max_scan + 1):
        if _norm(ws.cell(row=r, column=1).value) == expected_first:
            return r
    return None

def read_headers(ws, header_row, n):
    return [_norm(ws.cell(row=header_row, column=c).value) for c in range(1, n + 1)]

def declared_schema_version(wb):
    """Schema version is embedded in document Keywords as 'schema:X.Y'."""
    kw = _norm(getattr(wb.properties, "keywords", "") or "")
    m = re.search(r"schema:\s*([0-9]+\.[0-9]+)", kw)
    return m.group(1) if m else None


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------

def validate(path):
    rep = Report(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        rep.fatal("LOAD", f"Cannot open workbook: {e}")
        return rep

    sheets = wb.sheetnames

    # ---- Schema version (governance) ----
    sv = declared_schema_version(wb)
    if sv is None:
        rep.blocking("SV-01", f"No schema version embedded. Expected document Keywords to contain 'schema:{SCHEMA_VERSION}'.")
    elif sv != SCHEMA_VERSION:
        rep.blocking("SV-02", f"Workbook declares schema {sv}; validator enforces {SCHEMA_VERSION}.")

    # ---- WV-01/03: structural sheet contract (v2.4: Master Summary + feature sheet(s);
    #      the Review Summary sheet was removed at stakeholder request) ----
    if not sheets or sheets[0] != "Master Summary":
        rep.fatal("WV-01", f"First sheet must be 'Master Summary' (found {sheets[0] if sheets else 'none'!r}).")
    feature_sheets = [s for s in sheets[1:] if s != "Review Summary"]
    if len(feature_sheets) < 1:
        rep.fatal("WV-03", "At least one feature worksheet is required after Master Summary.")

    # ---- Per feature sheet ----
    seen_ids = {}  # id -> sheet
    for sname in feature_sheets:
        ws = wb[sname]

        # sheet name rules (WV-06)
        if len(sname) > 31:
            rep.blocking("WV-06", f"Sheet name {sname!r} exceeds 31 chars.")
        if any(ch in sname for ch in r'\/*?:[]'):
            rep.blocking("WV-06", f"Sheet name {sname!r} contains prohibited characters.")

        hrow = find_header_row(ws)
        if hrow is None:
            rep.fatal("WV-05", f"[{sname}] Could not locate a header row starting with 'Test Case ID'.")
            continue

        # WV-04/05: exact headers, exact width, no extra columns
        got = read_headers(ws, hrow, len(FEATURE_HEADERS))
        if got != FEATURE_HEADERS:
            rep.fatal("WV-05", f"[{sname}] Header mismatch.\n        expected: {FEATURE_HEADERS}\n        found:    {got}")
        extra = _norm(ws.cell(row=hrow, column=len(FEATURE_HEADERS) + 1).value)
        if extra:
            rep.fatal("WV-04", f"[{sname}] Extra column beyond canonical schema: {extra!r} (execution-tracking columns belong in Zephyr, not the workbook).")

        # ---- data rows ----
        _validate_feature_rows(ws, sname, hrow, seen_ids, rep)

    # ---- CV-01: Master Summary TC counts vs actual ----
    _validate_master_summary_counts(wb, feature_sheets, rep)

    # ---- CV-06/07: coverage completeness (verified from the Master Summary; no extra sheet) ----
    _validate_coverage_completeness(wb, rep)

    # ---- CV-08/09/10/11: verify coverage against the source-anchored coverage ledger sidecar ----
    _validate_coverage_ledger(path, wb, seen_ids, rep)

    # ---- NS-01/02: cross-workbook ID uniqueness + project namespacing ----
    _validate_namespacing(path, seen_ids, rep)

    # ---- DP-01/02: data-handling — no real PII / secrets in cells ----
    _validate_data_handling(wb, rep)

    # ---- DUP-01/EI-01: advisory semantic checks (near-dup + hard-coded env IDs) ----
    _validate_semantics(wb, rep)

    # ---- ST-02: advisory suite-level under-decomposition check ----
    _validate_decomposition(wb, rep)

    # NOTE (v2.4): the Review Summary sheet and its Requirement Traceability Matrix were
    # removed at stakeholder request, so RTM-based reverse-traceability / orphan detection
    # (RT-01..04) and coverage-recompute no longer run. Traceability is now a design-time
    # concern only (VALIDATION_ENGINE.md), not machine-enforced from the workbook.

    return rep


def _validate_coverage_completeness(wb, rep):
    """CV-06/07 — verify coverage completeness from the Master Summary (no extra columns or
    sheet). Enforces that the delivered suite claims COMPLETE AC coverage and is internally
    consistent, so an incomplete-coverage suite cannot be shipped silently.

    Honest limit: this proves coverage of the ACs the run *extracted*. Whether extraction
    itself captured every AC in the source cannot be verified from the workbook alone — that
    remains the design-time extraction-fidelity concern in VALIDATION_ENGINE.md."""
    if "Master Summary" not in wb.sheetnames:
        rep.blocking("CV-06", "No Master Summary sheet — coverage completeness cannot be verified.")
        return
    ms = wb["Master Summary"]
    hrow = None; col = {}
    for r in range(1, 8):
        names = {c: _norm(ms.cell(row=r, column=c).value).lower() for c in range(1, ms.max_column + 1)}
        if any(v == "test cases" for v in names.values()) and any("coverage" in v for v in names.values()):
            hrow = r
            for c, v in names.items():
                if v == "acceptance criteria": col["ac"] = c
                elif v == "test cases": col["tc"] = c
                elif "coverage" in v: col["cov"] = c
            break
    if hrow is None or "cov" not in col:
        rep.blocking("CV-06", "Could not locate the Master Summary 'AC Coverage %' column — coverage completeness cannot be verified.")
        return

    def _num(r, c):
        m = re.search(r"[0-9]+(?:\.[0-9]+)?", _norm(ms.cell(row=r, column=c).value))
        return float(m.group()) if m else None

    saw = False
    for r in range(hrow + 1, ms.max_row + 1):
        label = _norm(ms.cell(row=r, column=1).value)
        if not label:
            continue
        cov = _num(r, col["cov"])
        if cov is None:
            rep.blocking("CV-06", f"Master Summary row {r} ({label}): AC Coverage % is blank — completeness cannot be confirmed.")
        elif cov < 99.5:
            rep.fatal("CV-06", f"Master Summary row {r} ({label}): AC Coverage % = {cov:g}% — an incomplete-coverage suite must not be delivered (every AC needs >=1 test case).")
        if label.upper() == "TOTAL":
            continue
        saw = True
        ac = _num(r, col["ac"]) if "ac" in col else None
        tc = _num(r, col["tc"]) if "tc" in col else None
        if tc is not None and tc < 1:
            rep.blocking("CV-07", f"Master Summary row {r} ({label}): 0 test cases recorded.")
        if tc and tc >= 1 and ac is not None and ac < 1:
            rep.blocking("CV-07", f"Master Summary row {r} ({label}): {tc:g} test cases but 0 acceptance criteria — coverage cannot be assessed.")
    if not saw:
        rep.blocking("CV-06", "Master Summary has no per-feature data rows — coverage completeness cannot be verified.")


def _ms_header(ms):
    """Locate the Master Summary header row and the AC / Test Cases / Coverage columns."""
    for r in range(1, 8):
        names = {c: _norm(ms.cell(row=r, column=c).value).lower() for c in range(1, ms.max_column + 1)}
        if any(v == "test cases" for v in names.values()) and any("coverage" in v for v in names.values()):
            col = {}
            for c, v in names.items():
                if v == "acceptance criteria": col["ac"] = c
                elif v == "test cases":        col["tc"] = c
                elif "coverage" in v:          col["cov"] = c
            return r, col
    return None, {}

def _ms_totals(ms):
    """Total acceptance-criteria count and overall coverage % from the Master Summary
    (prefers an explicit TOTAL row; otherwise sums/averages the per-feature rows)."""
    hrow, col = _ms_header(ms)
    if hrow is None:
        return None
    def num(r, c):
        m = re.search(r"[0-9]+(?:\.[0-9]+)?", _norm(ms.cell(row=r, column=c).value))
        return float(m.group()) if m else None
    sum_ac = 0.0; cov_pcts = []; total_ac = None; total_cov = None
    for r in range(hrow + 1, ms.max_row + 1):
        label = _norm(ms.cell(row=r, column=1).value)
        if not label:
            continue
        ac  = num(r, col["ac"])  if "ac"  in col else None
        cov = num(r, col["cov"]) if "cov" in col else None
        if label.upper() == "TOTAL":
            total_ac, total_cov = ac, cov
            continue
        if ac is not None:  sum_ac += ac
        if cov is not None: cov_pcts.append(cov)
    return {
        "ac":  total_ac  if total_ac  is not None else sum_ac,
        "cov": total_cov if total_cov is not None else (sum(cov_pcts) / len(cov_pcts) if cov_pcts else None),
    }

def _ledger_path(xlsx_path):
    return re.sub(r"\.xlsx$", ".coverage.json", xlsx_path, flags=re.I)

def _feature_sheets_ordered(wb):
    return [s for s in wb.sheetnames[1:] if s != "Review Summary"]

def _feature_sheet_ids(wb):
    """Map each feature sheet -> the set of Test Case IDs that live ON that sheet."""
    out = {}
    for sname in _feature_sheets_ordered(wb):
        ws = wb[sname]
        hrow = find_header_row(ws)
        ids = set()
        if hrow is not None:
            for r in range(hrow + 1, ws.max_row + 1):
                v = _norm(ws.cell(row=r, column=1).value)
                if v:
                    ids.add(v)
        out[sname] = ids
    return out

def _ms_feature_rows(ms):
    """Ordered (label, ac_count) for each per-feature Master Summary row, excluding TOTAL."""
    hrow, col = _ms_header(ms)
    rows = []
    if hrow is None:
        return rows
    def num(r, c):
        m = re.search(r"[0-9]+(?:\.[0-9]+)?", _norm(ms.cell(row=r, column=c).value))
        return float(m.group()) if m else None
    for r in range(hrow + 1, ms.max_row + 1):
        label = _norm(ms.cell(row=r, column=1).value)
        if not label or label.upper() == "TOTAL":
            continue
        rows.append((label, num(r, col["ac"]) if "ac" in col else None))
    return rows

def _validate_coverage_ledger(path, wb, seen_ids, rep):
    """CV-08/09/10/11 — verify AC coverage against a source-anchored coverage ledger sidecar
    (`<workbook>.coverage.json`), converting the self-reported Master Summary percentage into a
    checkable, itemized AC-to-TC mapping. Coverage is verified PER FEATURE: an AC's covering
    test cases must live on that AC's own feature sheet (no cross-sheet borrowing), and each
    feature's AC count must reconcile with its Master Summary row. Honest limit: the ledger
    proves every AC it *declares* is genuinely covered by real test cases on its own feature and
    cites a source anchor; it still cannot prove the ledger enumerated every AC in the source —
    that remains a design-time extraction-fidelity + human sampling concern."""
    lp = _ledger_path(path)
    if not os.path.exists(lp):
        rep.blocking("CV-11", f"No coverage ledger sidecar found ({os.path.basename(lp)}). AC coverage cannot be verified against a source-anchored AC list — author the ledger during extraction and deliver it alongside the workbook.")
        return
    try:
        led = json.load(open(lp, encoding="utf-8"))
    except Exception as e:
        rep.blocking("CV-11", f"Coverage ledger {os.path.basename(lp)} is not valid JSON: {e}")
        return
    feats = led.get("features")
    if not isinstance(feats, list) or not feats:
        rep.fatal("CV-08", f"Coverage ledger {os.path.basename(lp)} has no non-empty 'features' list.")
        return

    per_sheet = _feature_sheet_ids(wb)
    feature_sheets = _feature_sheets_ordered(wb)
    ms_rows = _ms_feature_rows(wb["Master Summary"]) if "Master Summary" in wb.sheetnames else []

    ledger_ac = 0
    ledger_covered = 0
    for feat in feats:
        sheet = _norm(feat.get("sheet")) or "<unnamed feature>"
        acs = feat.get("acceptance_criteria") or []
        if not acs:
            rep.fatal("CV-08", f"Ledger feature {sheet!r} declares no acceptance criteria.")
            continue
        if sheet in per_sheet:
            scope = per_sheet[sheet]          # per-feature: only this sheet's TC IDs count as coverage
        else:
            rep.blocking("CV-08", f"Ledger feature {sheet!r} does not match any feature sheet in the workbook.")
            scope = seen_ids                  # degrade gracefully: still check the TCs exist somewhere
        feat_ac = 0
        for ac in acs:
            ledger_ac += 1
            feat_ac += 1
            acid   = _norm(ac.get("id")) or "?"
            anchor = _norm(ac.get("anchor"))
            cov    = ac.get("covered_by") or []
            if not anchor:
                rep.blocking("CV-10", f"Ledger AC {acid} in {sheet!r} has no source anchor — every AC must cite where in the source it came from.")
            not_in_wb   = [t for t in cov if t not in seen_ids]
            wrong_sheet = [t for t in cov if t in seen_ids and t not in scope]
            in_scope    = [t for t in cov if t in scope]
            if not_in_wb:
                rep.fatal("CV-08", f"Ledger AC {acid} in {sheet!r} lists covering test case(s) not present in the workbook: {not_in_wb}.")
            if wrong_sheet:
                rep.fatal("CV-08", f"Ledger AC {acid} in {sheet!r} is covered by test case(s) that belong to a different feature sheet: {wrong_sheet} — an AC must be covered by test cases on its own feature.")
            if in_scope:
                ledger_covered += 1
            elif not not_in_wb and not wrong_sheet:
                rep.fatal("CV-08", f"Ledger AC {acid} in {sheet!r} has no existing covering test case — the acceptance criterion is uncovered.")
        # CV-09 per-feature: ledger AC count for this feature must match its Master Summary row
        if sheet in feature_sheets:
            idx = feature_sheets.index(sheet)
            if idx < len(ms_rows):
                ms_label, ms_ac = ms_rows[idx]
                if ms_ac is not None and abs(ms_ac - feat_ac) > 0.5:
                    rep.fatal("CV-09", f"Feature {sheet!r}: Master Summary row ({ms_label}) reports {ms_ac:g} acceptance criteria but the ledger itemizes {feat_ac} for this feature.")

    # CV-09 totals backstop: reconcile ledger grand totals against the Master Summary TOTAL row.
    if "Master Summary" in wb.sheetnames and ledger_ac:
        tot = _ms_totals(wb["Master Summary"])
        if tot:
            if tot["ac"] is not None and abs(tot["ac"] - ledger_ac) > 0.5:
                rep.fatal("CV-09", f"Master Summary reports {tot['ac']:g} acceptance criteria in total but the ledger itemizes {ledger_ac} — the reported AC count must match the source-anchored list.")
            recomputed = round(100.0 * ledger_covered / ledger_ac, 1)
            if tot["cov"] is not None and abs(tot["cov"] - recomputed) > 0.6:
                rep.fatal("CV-09", f"Master Summary AC Coverage % ({tot['cov']:g}%) does not match coverage recomputed from the ledger ({recomputed:g}%) — the reported number must equal the real AC-to-TC mapping.")


# --------------------------------------------------------------------------
# Multi-project namespacing + persistent global ID ledger
# --------------------------------------------------------------------------
# Two operational-state files live beside this validator:
#   project_registry.json  — declares the known project keys / business units (namespacing).
#   id_ledger.json         — the persistent record of every TC ID ever issued, so that reuse
#                            of an ID across DIFFERENT workbooks (a real collision that the
#                            per-workbook DV-09 check cannot see) is detected. IDs are added to
#                            the ledger only via `--register` after a workbook passes.
_HERE = os.path.dirname(os.path.abspath(__file__))

def _load_beside(name):
    p = os.path.join(_HERE, name)
    if not os.path.exists(p):
        return None
    try:
        return json.load(open(p, encoding="utf-8"))
    except Exception:
        return None

def _project_key(tc_id):
    m = re.match(r"([A-Z][A-Z0-9]*)-\d+-TC-\d+", tc_id)
    return m.group(1) if m else None

def _validate_namespacing(path, seen_ids, rep):
    """NS-01 (cross-workbook ID collision) + NS-02 (unregistered project key)."""
    wbname = os.path.basename(path)
    reg = _load_beside("project_registry.json")
    if reg:
        enabled = {p.get("key") for p in reg.get("projects", []) if p.get("enabled", True)}
        seen_keys = {k for k in (_project_key(i) for i in seen_ids) if k}
        for k in sorted(seen_keys - enabled):
            rep.warn("NS-02", f"Project key {k!r} is not an enabled project in project_registry.json — register the project so its IDs are namespaced and governed.")
    led = _load_beside("id_ledger.json")
    if led:
        issued = led.get("issued", {})
        for tid in sorted(seen_ids):
            rec = issued.get(tid)
            if rec and rec.get("workbook") != wbname:
                rep.fatal("NS-01", f"Test Case ID {tid!r} was already issued in a different workbook ({rec.get('workbook')!r}) — cross-workbook ID collision.")

def register_ids(path):
    """Append a PASSED workbook's TC IDs to the persistent id_ledger.json. Returns (added, msg)."""
    rep = validate(path)
    if rep.failed:
        return 0, f"NOT registered — workbook failed validation: {path}"
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return 0, f"NOT registered — cannot open {path}: {e}"
    ids = set()
    for sname in [s for s in wb.sheetnames[1:] if s != "Review Summary"]:
        ws = wb[sname]
        hrow = find_header_row(ws)
        if hrow is None:
            continue
        for r in range(hrow + 1, ws.max_row + 1):
            v = _norm(ws.cell(row=r, column=1).value)
            if v:
                ids.add(v)
    led = _load_beside("id_ledger.json") or {"schema": "id-ledger-1.0", "issued": {}}
    issued = led.setdefault("issued", {})
    wbname = os.path.basename(path)
    today = date.today().isoformat()
    added = 0
    for i in sorted(ids):
        if i not in issued:
            issued[i] = {"workbook": wbname, "date": today}
            added += 1
        elif issued[i].get("workbook") != wbname:
            return 0, f"NOT registered — {i} already issued in {issued[i]['workbook']!r} (collision)"
    json.dump(led, open(os.path.join(_HERE, "id_ledger.json"), "w", encoding="utf-8"), indent=1)
    return added, f"registered {added} new ID(s) from {wbname} (ledger now holds {len(issued)})"


# --------------------------------------------------------------------------
# Data-handling enforcement (DATA_HANDLING.md) — high-precision PII / secret scan
# --------------------------------------------------------------------------
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_SECRET_RES = [
    re.compile(r"AKIA[0-9A-Z]{16}"),                                  # AWS access key id
    re.compile(r"-----BEGIN [A-Z ]*PRIVATE KEY-----"),               # PEM private key
    re.compile(r"\bghp_[A-Za-z0-9]{20,}\b"),                          # GitHub token
    re.compile(r"\bxox[baprs]-[A-Za-z0-9-]{10,}"),                    # Slack token
    re.compile(r"\beyJ[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}\b"),  # JWT
]

def _validate_data_handling(wb, rep):
    """DP-01/DP-02 — reject real PII (email) or secrets/credentials reproduced in cells.
    Deliberately high-precision (literal emails and well-known secret formats) so that
    characteristic-first placeholders like `<email>` or 'the approver's email' never trip it."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                if _EMAIL_RE.search(v):
                    rep.blocking("DP-01", f"[{ws.title}] {c.coordinate}: contains a real-looking email address — use an environment-independent placeholder/characteristic, not real personal data.")
                for rx in _SECRET_RES:
                    if rx.search(v):
                        rep.blocking("DP-02", f"[{ws.title}] {c.coordinate}: contains a probable secret/credential — secrets must never be reproduced in test cases.")
                        break


# --------------------------------------------------------------------------
# Cheap semantic checks (advisory): near-duplicate scenarios + hard-coded env IDs
# --------------------------------------------------------------------------
# Probable hard-coded environment/record identifiers (high-precision to avoid false positives).
_ENVID_RES = [
    re.compile(r"\bQ-\d{4,}\b"),                                  # quote id e.g. Q-100245
    re.compile(r"\b(?:001|003|005|006|00Q|500|701|807)[A-Za-z0-9]{12}(?:[A-Za-z0-9]{3})?\b"),  # Salesforce 15/18-char id
    re.compile(r"\b(?:OPP|ACC|QT|CASE)-\d{3,}\b"),               # generic record ids
]

def _sig(text):
    return re.sub(r"\s+", " ", str(text).strip().lower())

# ST-01: a Test Step that bundles two UI actions (an action verb, a connector, then a SECOND
# action verb) is a compound step and should be split into one action per step. High-precision:
# the connector must be followed by another action verb, so noun-phrase "and"s
# ("Notes and Attachments", "reminders, expiry and email body text") never trip it.
_ACTION_VERB = (r"(?:click|enter|select|choose|navigate|open|set|save|check|uncheck|add|remove|"
                r"apply|submit|press|download|upload|confirm|trigger|recall|send|generate|assign|"
                r"place|reopen|create|edit|approve|reinitiate|configure|initiate|mark|attach|fill|"
                r"toggle|activate|deactivate|delete|complete|decline|preview|reject)")
_COMPOUND_STEP_RE = re.compile(r"(?:,?\s+then\s+|\s+and\s+then\s+|\s+and\s+)" + _ACTION_VERB + r"\b", re.I)

def _validate_semantics(wb, rep):
    """DUP-01 (near-duplicate test cases, per feature) + EI-01 (hard-coded env/record IDs)."""
    for sname in _feature_sheets_ordered(wb):
        ws = wb[sname]
        hrow = find_header_row(ws)
        if hrow is None:
            continue
        # gather per-TC title + (step, expected) signature
        order = []
        tcs = {}
        cur = None
        for r in range(hrow + 1, ws.max_row + 1):
            a = _norm(ws.cell(row=r, column=1).value)
            if a and a != cur:
                cur = a
                order.append(cur)
                tcs[cur] = {"title": _norm(ws.cell(row=r, column=3).value), "sig": []}
            if cur is None:
                continue
            st = _norm(ws.cell(row=r, column=6).value)
            ex = _norm(ws.cell(row=r, column=7).value)
            if st or ex:
                tcs[cur]["sig"].append((_sig(st), _sig(ex)))
            # ST-01: flag compound (multi-action) test steps
            if st:
                m = _COMPOUND_STEP_RE.search(st)
                if m:
                    rep.warn("ST-01", f"[{sname}] {ws.cell(row=r, column=6).coordinate} ({cur}): test step appears to bundle multiple actions ({m.group().strip()!r}) — split into one UI action per step (TEST_CASE_GENERATION.md 5.2).")
            # EI-01: scan Pre-Conditions (D), Test Step (F), Expected Result (G)
            for col in (4, 6, 7):
                v = _norm(ws.cell(row=r, column=col).value)
                for rx in _ENVID_RES:
                    m = rx.search(v)
                    if m:
                        rep.warn("EI-01", f"[{sname}] {ws.cell(row=r, column=col).coordinate}: probable hard-coded environment/record identifier ({m.group()!r}) — use an environment-independent characteristic instead.")
                        break
        # DUP-01: identical normalized title OR identical step+expected signature
        seen_title = {}
        seen_sig = {}
        for tid in order:
            t = _sig(tcs[tid]["title"])
            s = tuple(tcs[tid]["sig"])
            if t and t in seen_title:
                rep.warn("DUP-01", f"[{sname}] {tid} and {seen_title[t]} have identical titles — probable duplicate scenario; merge or differentiate.")
            elif t:
                seen_title[t] = tid
            if s and s in seen_sig:
                rep.warn("DUP-01", f"[{sname}] {tid} and {seen_sig[s]} have identical steps and expected results — probable duplicate; merge or differentiate.")
            elif s:
                seen_sig[s] = tid


def _validate_decomposition(wb, rep):
    """ST-02 (advisory, suite-level) — flag an under-decomposed suite: too high a share of
    test cases sit at exactly the 3-step floor. TEST_CASE_GENERATION.md 5.1 expects most
    manual cases to run 4-8 atomic steps (one UI action per step) and warns that a suite
    where most cases are exactly three steps is almost certainly under-decomposed. This turns
    that prose guidance into a machine check. Emitted once for the whole workbook; only
    evaluated on a suite large enough (>=8 TCs) for the ratio to be meaningful."""
    three = 0
    total = 0
    for sname in _feature_sheets_ordered(wb):
        ws = wb[sname]
        hrow = find_header_row(ws)
        if hrow is None:
            continue
        cur = None
        steps = 0
        for r in range(hrow + 1, ws.max_row + 1):
            a = _norm(ws.cell(row=r, column=1).value)
            st = _norm(ws.cell(row=r, column=6).value)
            if a and a != cur:            # a new test case begins (merged/forward-filled)
                if cur is not None:
                    total += 1
                    if steps == 3:
                        three += 1
                cur = a
                steps = 0
            if st:
                steps += 1
        if cur is not None:
            total += 1
            if steps == 3:
                three += 1
    THRESHOLD = 0.60
    if total >= 8:
        share = three / total
        if share > THRESHOLD:
            rep.warn("ST-02", f"{three} of {total} test cases ({share*100:.0f}%) have exactly 3 steps — the suite looks under-decomposed. Most manual cases should run 4-8 atomic steps (one UI action per step); only genuinely simple presence/absence checks belong at the 3-step floor (TEST_CASE_GENERATION.md 5.1/5.2).")


def _validate_feature_rows(ws, sname, hrow, seen_ids, rep):
    """Walk step rows. v2.3: TC-level fields (Test Case ID, Requirement Title, Test Case
    Title, Pre-Conditions, Priority) are MERGED vertically across a test case's step rows,
    so the value appears only on the first row and continuation rows read blank. A new test
    case begins on a row where Test Case ID is populated. (Forward-filled workbooks, where
    the value repeats on every row, are also tolerated — a repeated identical ID is treated
    as a continuation, not a new TC.)"""
    col = {h: i + 1 for i, h in enumerate(FEATURE_HEADERS)}
    current_id = None
    steps_in_tc = 0
    expected_step = 0
    last_expected = ""      # Expected Result of the most recent step (the TC's final step at close)

    def close_tc():
        if current_id is not None and steps_in_tc < MIN_STEPS_PER_TC:
            rep.warn("DV-10", f"[{sname}] {current_id} has {steps_in_tc} step(s); minimum is {MIN_STEPS_PER_TC}.")
        # ER-02: the final (key verification) step should carry the full set of checks.
        if current_id is not None and last_expected and not _is_banned_expected(last_expected):
            lines = _expected_lines(last_expected)
            if len(lines) == 1 and len(lines[0]) < _THIN_FINAL_MAX:
                rep.warn("ER-02", f"[{sname}] {current_id}: final-step Expected Result looks thin ({lines[0]!r}); the key verification step should state the full set of observable checks (TEST_CASE_GENERATION.md §6.1).")

    r = hrow + 1
    max_row = ws.max_row
    while r <= max_row:
        tc_id = _norm(ws.cell(row=r, column=col["Test Case ID"]).value)
        step_v = ws.cell(row=r, column=col["Step#"]).value
        test_step = _norm(ws.cell(row=r, column=col["Test Step"]).value)
        expected = _norm(ws.cell(row=r, column=col["Expected Result"]).value)

        # skip fully-blank trailing rows
        if not tc_id and step_v is None and not test_step and not expected:
            r += 1
            continue

        # New TC when a NEW id appears; blank id (merged) or repeated id = continuation.
        is_new_tc = bool(tc_id) and tc_id != current_id
        if not tc_id and current_id is None:
            rep.fatal("DV-01", f"[{sname}] row {r}: a step row appears before any Test Case ID (first TC row must carry the ID).")
            r += 1
            continue

        if is_new_tc:
            close_tc()
            current_id = tc_id
            steps_in_tc = 0
            expected_step = 0
            last_expected = ""

            if not TC_ID_RE.match(tc_id):
                rep.fatal("DV-11", f"[{sname}] row {r}: TC ID {tc_id!r} does not match required format PROJECT-STORY-TC-NNN (e.g. SAMP-125-TC-001).")
            if tc_id in seen_ids:
                rep.fatal("DV-09", f"[{sname}] Test Case ID {tc_id!r} reappears as a NEW test case (also in {seen_ids[tc_id]}). IDs must be globally unique.")
            else:
                seen_ids[tc_id] = sname

            # TC-level fields live on the first (merge-anchor) row
            title = _norm(ws.cell(row=r, column=col["Test Case Title"]).value)
            if not title:
                rep.fatal("DV-02", f"[{sname}] {tc_id}: blank Test Case Title.")
            elif not title.startswith(TITLE_PREFIXES):
                rep.warn("DV-08", f"[{sname}] {tc_id}: title should start with {TITLE_PREFIXES}.")
            if not _norm(ws.cell(row=r, column=col["Requirement Title"]).value):
                rep.blocking("DV-12", f"[{sname}] {tc_id}: blank Requirement Title.")
            if not _norm(ws.cell(row=r, column=col["Pre-Conditions"]).value):
                rep.blocking("DV-13", f"[{sname}] {tc_id}: blank Pre-Conditions.")
            prio = _norm(ws.cell(row=r, column=col["Priority"]).value)
            if not prio:
                rep.blocking("DV-05", f"[{sname}] {tc_id}: blank Priority.")
            elif prio not in PRIORITY_VALUES:
                rep.blocking("DV-06", f"[{sname}] {tc_id}: priority {prio!r} not in {sorted(PRIORITY_VALUES)}.")

        # every step row (first or continuation):
        steps_in_tc += 1
        expected_step += 1
        if not test_step:
            rep.fatal("DV-03", f"[{sname}] row {r} ({current_id}): blank Test Step.")
        if not expected:
            rep.fatal("DV-04", f"[{sname}] row {r} ({current_id}): blank Expected Result.")
        elif _is_banned_expected(expected):
            rep.blocking("ER-01", f"[{sname}] row {r} ({current_id}): Expected Result {expected!r} is a vague/bare outcome — state the specific observable result (TEST_CASE_GENERATION.md §6.4).")
        last_expected = expected
        try:
            if step_v is not None and int(step_v) != expected_step:
                rep.warn("DV-07", f"[{sname}] {current_id}: Step# {step_v} out of sequence (expected {expected_step}).")
        except (ValueError, TypeError):
            rep.warn("DV-07", f"[{sname}] row {r} ({current_id}): Step# {step_v!r} is not an integer.")
        r += 1

    close_tc()


def _validate_master_summary_counts(wb, feature_sheets, rep):
    """CV-01: the Test Cases count in Master Summary must match the actual
    number of test cases on each feature sheet. Deterministic, no LLM claim trusted."""
    # actual counts
    actual = {}
    for sname in feature_sheets:
        ws = wb[sname]
        hrow = find_header_row(ws)
        if hrow is None:
            continue
        # count DISTINCT test case IDs (forward-fill repeats the ID on every step row)
        distinct = set()
        for r in range(hrow + 1, ws.max_row + 1):
            v = _norm(ws.cell(row=r, column=1).value)
            if v:
                distinct.add(v)
        actual[sname] = len(distinct)

    ms = wb["Master Summary"] if "Master Summary" in wb.sheetnames else None
    if ms is None:
        return
    # locate the Test Cases column in the Master Summary header row
    ms_hrow = None
    tc_col = None
    for r in range(1, 8):
        for c in range(1, ms.max_column + 1):
            if _norm(ms.cell(row=r, column=c).value) == "Test Cases":
                ms_hrow, tc_col = r, c
                break
        if ms_hrow:
            break
    if ms_hrow is None:
        rep.warn("CV-01", "Could not locate 'Test Cases' column in Master Summary; skipped count cross-check.")
        return

    total_claimed = 0
    total_actual = sum(actual.values())
    for r in range(ms_hrow + 1, ms.max_row + 1):
        label = _norm(ms.cell(row=r, column=1).value)
        val = ms.cell(row=r, column=tc_col).value
        if not label or label.upper() == "TOTAL":
            continue
        try:
            claimed = int(val)
        except (ValueError, TypeError):
            continue
        total_claimed += claimed
    if total_claimed and total_claimed != total_actual:
        rep.fatal("CV-01", f"Master Summary claims {total_claimed} test cases; feature sheets actually contain {total_actual}.")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main(argv):
    if "--rules" in argv:
        print(emit_rules_block())
        probs = rules_consistency_errors()
        for p in probs:
            sys.stderr.write(f"RULES DRIFT: {p}\n")
        return 1 if probs else 0
    if "--register" in argv:
        targets = [a for a in argv if not a.startswith("--")]
        if not targets:
            sys.stderr.write("--register requires at least one workbook path\n")
            return 2
        rc = 0
        for p in targets:
            added, msg = register_ids(p)
            print(msg)
            if "NOT registered" in msg:
                rc = 1
        return rc
    # default: validate the given workbook(s) via the shared CLI (paths + --json)
    return run_cli(validate, argv, __doc__)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
